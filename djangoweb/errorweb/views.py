from django.shortcuts import render 
from django.http import HttpResponse
from openpyxl.workbook.workbook import Workbook  
from errorweb.functions import handle_uploaded_file  
from djangoweb.forms import StudentForm  
from django.contrib import messages

def home(request):
    return render(request,"home.html")

def index(request):  
    if request.method == 'POST':  
        student = StudentForm(request.POST, request.FILES)  
        if student.is_valid():  
            f1=request.FILES['file'].name
            f2=request.FILES['file2'].name
            if f1.endswith('.xlsx'):
                handle_uploaded_file(request.FILES['file2'])
            else:
                messages.add_message(request,messages.INFO,f1 +" file is invalid for operation")
            if f2.endswith('.xlsx'):
                handle_uploaded_file(request.FILES['file']) 
            else:
                messages.add_message(request,messages.INFO,f2 +" file is invalid for operation")
            if f1.endswith('.xlsx') and f2.endswith('.xlsx'):
                execute(f1,f2)
                return HttpResponse("downloaded successfully")
                return render(request,'download.html')
            else:
                return render(request,'index.html',{'form':student})
    else:  
        student = StudentForm()  
        return render(request,"index.html",{'form':student}) 

def execute(f1,f2):
    import openpyxl
    from openpyxl.workbook.workbook import Workbook

    #folder1="D:\\ALIMweb\\djangoweb\\upload\\error\\"
    #folder2="D:\\ALIMweb\\djangoweb\\upload\\validation\\"

    #name="PROD_TOL_OPS01_Register_Batch26_Part2"
    wb=openpyxl.load_workbook("D:\\ALIMweb\\djangoweb\\upload\\error\\"+f1)
    loadsheet=wb.active

    wb1=openpyxl.load_workbook("D:\\ALIMweb\\djangoweb\\upload\\validation\\"+f2)
    valsheet=wb1[wb1.sheetnames[1]]

    output=Workbook()
    TOL=output.active

    for i in range(1,valsheet.max_column+1):
        TOL.cell(column=i,row=1).value=valsheet.cell(column=i,row=1).value
    #output.save("PROD_CAS_PROC03.xlsx")


    for i in range(1,loadsheet.max_row+1):
    #   print(loadsheet.cell(column=1,row=i).value)
        if loadsheet.cell(column=1,row=i).value!=None and loadsheet.cell(column=1,row=i).value.find('Row:')!=-1:
            rownum=int(loadsheet.cell(column=1,row=i).value.partition('Row:')[2].partition(' ')[0])
            Sheetname=loadsheet.cell(column=1,row=i).value.partition('Sheet:')[2].partition(' ')[0].replace("'",'')
            Sheetname=wb1[Sheetname]
            l=[]
            for k in range(1,Sheetname.max_column+1):
                l.append(Sheetname.cell(row=rownum,column=k).value)
            t=tuple(l)
            TOL.append(t)
    filename=f2+"_Import_Error.xlsx"
    output.save("D:\\ALIMweb\\djangoweb\\upload\\output\\"+filename)


